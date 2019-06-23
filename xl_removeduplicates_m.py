# https://github.com/love2spooge/python_xl_removeduplicates
# Removing duplicates based on multiple columns
# Change col_no= add multiple columns (column A=1, B=2 etc...), keep one for single column
# If you data does not have headers change start_row = 1 to 0
# start_row also defines starting row

# /// import + other system stuff
import openpyxl
from openpyxl.utils.cell import get_column_letter

# /// variables
wb = openpyxl.load_workbook("sample_s.xlsx") # open file
sheet = wb.active # data sheet (first sheet)

values = []
start_row = 1
col_no = [1, 2]
row_string = ""

#CODE

for i in range(1, sheet.max_row):

    for x in col_no:
        row_string = row_string + str(sheet.cell(row=start_row,column=x).value)

    values.insert(0, row_string)

    if "None" in values: # if cell is empty it is skiped
        values.remove("None")

    row_string = ""
    for x in col_no:
        row_string = row_string + str(sheet.cell(row=start_row+1,column=x).value)

    if row_string in values:
        print("Duplicate found, removing row:", (start_row+1))
        sheet.delete_rows((start_row+1), amount=1)
        start_row = start_row - 1 # this is a must as each time when row is removed you need to check previous row also

    start_row = start_row + 1
    row_string = ""

wb.save("sample_m_done.xlsx")
