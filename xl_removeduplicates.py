# https://github.com/love2spooge/python_xl_removeduplicates

# /// import + other system stuff
import sys
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils import column_index_from_string

# /// variables
col_no = []
values = set()
start_row = 1 # 1 = my data has headers
row_string = ""
counter = 0

xl_file = sys.argv[1] if len(sys.argv) > 1 else "" # argument from command line

if xl_file == "": # if argument is not specified
    print(
    '''
    You didn't specified the input file
    Make sure that you run script with argument:
    python xl_removeduplicates.py sample.xlsx
    '''
    )
    quit()

xl_columns = sys.argv[2] if len(sys.argv) > 2 else "" # argument from command line

if xl_columns == "": # if argument is not specified
    print(
    '''
    You didn't specified columns(s)
    Make sure that you run script with argument:
    python xl_removeduplicates.py sample.xlsx A,B
    '''
    )
    quit()

xl_headers = sys.argv[3] if len(sys.argv) > 3 else "" # argument from command line
if xl_headers == "-h":
    start_row = 2 # 2 = my data has headers
else:
    start_row = 1


for i in xl_columns.split(","):
    col_no.insert(0,column_index_from_string(i))

wb = openpyxl.load_workbook(xl_file) # open file
sheet = wb.active # data sheet (first sheet)

#CODE

print(">> Opening file:", xl_file)
print(">> Checking duplicates in the following column(s):", sorted(xl_columns.split(",")))
if xl_headers == "-h":
    print(">> My data has headers")

for i in range(1, sheet.max_row):

    for x in sorted(col_no):
        row_string = row_string + str(sheet.cell(row=start_row,column=x).value)

    values.add(row_string) 

    if "None" in values: # if cell is empty it is skiped
        values.remove("None")

    row_string = ""
    for x in sorted(col_no):
        row_string = row_string + str(sheet.cell(row=start_row+1,column=x).value)
        

    if row_string in values:
        sheet.delete_rows((start_row+1), amount=1)
        start_row = start_row - 1 # this is a must as each time when row is removed you need to check previous row also
        counter = counter + 1

    start_row = start_row + 1
    row_string = ""

wb.save("removed_" + xl_file)

print("")
print(">> Done")
print(">>", counter, "duplicates found and", counter, "duplicates removed")
print(">> File saved:", "removed_" + xl_file)
