# https://github.com/love2spooge/python_xl_removeduplicates
# Removing duplicates based on single column
# Change col_no=2 for column (column A=1, B=2 etc...)
# If you data does not have headers change start_row = 1 to 0
# start_row also defines starting row

# /// import + other system stuff
import openpyxl
from openpyxl.utils.cell import get_column_letter

# /// variables
wb = openpyxl.load_workbook("sample_s.xlsx") # open file
sheet = wb.active # data sheet (first sheet)

values = set()
start_row = 1
col_no = 2

#CODE

for i in range(1, sheet.max_row):
    values.add(str(sheet.cell(row=start_row,column=col_no).value))

    if "None" in values: # if cell is empty it is skiped
        values.remove("None")

    print("")
    print(">Checking row:", start_row, "in column:", get_column_letter(col_no), "with value:", sheet.cell(row=(start_row+1), column=col_no).value, "and comparing it to:")
    print(values)

    if sheet.cell(row=start_row+1, column=col_no).value in values:
        print("Duplicate found, removing row:", (start_row+1))
        sheet.delete_rows((start_row+1), amount=1)
        start_row = start_row - 1 # this is a must as each time when row is removed you need to check previous row also

    start_row = start_row + 1

wb.save("sample_s_done.xlsx")
